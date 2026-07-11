"""File-type detection and CSV/XLSX parsing — entity-agnostic. Verifies
the actual content can be parsed rather than trusting the filename alone,
enforces file-size/row-count limits, and strips fully-empty rows.
"""

import csv
import io

from openpyxl import load_workbook

from app.imports.types import (
    REASON_EMPTY_FILE,
    REASON_FILE_TOO_LARGE,
    REASON_MALFORMED_FILE,
    REASON_TOO_MANY_ROWS,
    REASON_UNSUPPORTED_FILE_TYPE,
    ImportFileType,
    ParsedFile,
)

# ZIP local-file-header signature — every .xlsx (a zip archive) starts with
# this, regardless of what its filename claims.
_XLSX_MAGIC = b"PK\x03\x04"
# OLE2/CFB signature — legacy .xls (and old .doc/.ppt) files start with
# this. Detected specifically so a renamed/misnamed .xls gets a clear
# "not supported" message rather than a confusing parse failure.
_XLS_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class ImportFileError(Exception):
    """Raised with a stable reason_code (see app.imports.types) for any
    file-level problem. Routers catch this and map reason_code to the
    right HTTP status (415/400/413) — see app/routers/customer_imports.py."""

    def __init__(self, reason_code: str, message: str):
        self.reason_code = reason_code
        self.message = message
        super().__init__(message)


def _extension_of(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def detect_file_type(filename: str, content: bytes) -> ImportFileType:
    """Determines CSV vs XLSX from the actual bytes, not just the
    filename — an XLSX must have the ZIP signature; anything claiming to
    be XLSX without it is rejected outright rather than handed to
    openpyxl. Legacy .xls is explicitly rejected with its own message."""
    if content.startswith(_XLS_OLE2_MAGIC):
        raise ImportFileError(
            REASON_UNSUPPORTED_FILE_TYPE,
            "Legacy .xls files are not supported. Please save the file as .xlsx or .csv.",
        )

    ext = _extension_of(filename)

    if ext == "xlsx":
        if not content.startswith(_XLSX_MAGIC):
            raise ImportFileError(
                REASON_UNSUPPORTED_FILE_TYPE,
                "This file has a .xlsx extension but its content isn't a valid XLSX file.",
            )
        return ImportFileType.xlsx

    if content.startswith(_XLSX_MAGIC):
        # Content is unambiguously a zip/xlsx archive regardless of what
        # the extension says.
        return ImportFileType.xlsx

    if ext == "csv":
        return ImportFileType.csv

    if not ext:
        # No extension at all: fall back to sniffing whether it's plausibly
        # text (i.e. CSV), rather than rejecting solely for a missing
        # extension.
        try:
            content.decode("utf-8-sig")
            return ImportFileType.csv
        except UnicodeDecodeError:
            pass

    raise ImportFileError(
        REASON_UNSUPPORTED_FILE_TYPE,
        f"Unsupported file type ({('.' + ext) if ext else 'unknown'}). "
        "Please upload a .csv or .xlsx file.",
    )


def _row_is_empty(values) -> bool:
    return all(not str(v).strip() for v in values if v is not None) or not any(
        v is not None for v in values
    )


def parse_csv(content: bytes) -> ParsedFile:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportFileError(
            REASON_MALFORMED_FILE,
            "This CSV file couldn't be read as UTF-8 text.",
        ) from exc

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel  # comma-delimited default

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    try:
        all_rows = list(reader)
    except csv.Error as exc:
        raise ImportFileError(REASON_MALFORMED_FILE, "This CSV file couldn't be parsed.") from exc

    if not all_rows:
        raise ImportFileError(REASON_EMPTY_FILE, "This file has no rows.")

    headers = [h.strip() for h in all_rows[0]]
    data_rows: list[dict[str, str]] = []
    for raw_row in all_rows[1:]:
        if _row_is_empty(raw_row):
            continue
        # Rows shorter/longer than the header row are padded/truncated so
        # every declared header always has a (possibly empty) value.
        padded = list(raw_row) + [""] * (len(headers) - len(raw_row))
        data_rows.append({headers[i]: padded[i].strip() for i in range(len(headers))})

    return ParsedFile(file_type=ImportFileType.csv, headers=headers, rows=data_rows)


def parse_xlsx(content: bytes) -> ParsedFile:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        all_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    except Exception as exc:  # openpyxl raises various exceptions for corrupt files
        raise ImportFileError(REASON_MALFORMED_FILE, "This XLSX file couldn't be parsed.") from exc

    if not all_rows:
        raise ImportFileError(REASON_EMPTY_FILE, "This file has no rows.")

    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    data_rows: list[dict[str, str]] = []
    for raw_row in all_rows[1:]:
        if _row_is_empty(raw_row):
            continue
        padded = list(raw_row) + [None] * (len(headers) - len(raw_row))
        data_rows.append(
            {
                headers[i]: ("" if padded[i] is None else str(padded[i]).strip())
                for i in range(len(headers))
            }
        )

    return ParsedFile(file_type=ImportFileType.xlsx, headers=headers, rows=data_rows)


def parse_upload(
    filename: str,
    content: bytes,
    max_bytes: int,
    max_rows: int,
) -> ParsedFile:
    """The single entry point routers call: size check, type detection,
    parsing, row-count check, all in one place so no call site can skip a
    step."""
    if len(content) == 0:
        raise ImportFileError(REASON_EMPTY_FILE, "The uploaded file is empty.")
    if len(content) > max_bytes:
        raise ImportFileError(
            REASON_FILE_TOO_LARGE,
            f"The file is larger than the {max_bytes // 1_000_000} MB limit.",
        )

    file_type = detect_file_type(filename, content)
    parsed = parse_csv(content) if file_type == ImportFileType.csv else parse_xlsx(content)

    if not parsed.rows:
        raise ImportFileError(REASON_EMPTY_FILE, "This file has no data rows.")
    if len(parsed.rows) > max_rows:
        raise ImportFileError(
            REASON_TOO_MANY_ROWS,
            f"This file has more than the {max_rows}-row limit.",
        )

    return parsed
